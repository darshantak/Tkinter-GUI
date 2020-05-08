from tkinter import *
from openpyxl import *
from tkinter import messagebox
from xlrd import *
# import setuptools
# import distutils
# import site

worksheet=load_workbook('F:\workspace_python\Geeks For Geeks\excel1.xlsx')
sheet=worksheet.active

def clear():
    name_field.delete(0,END)
    guardian_field.delete(0,END)
    join_date_field.delete(0,END)
    end_date_field.delete(0,END)
    contact_field.delete(0,END)
    address_field.delete(0,END)
    age_field.delete(0,END)


def insert():
    if ( name_field.get()=="" or 
    guardian_field.get()=="" or 
    join_date_field.get()=="" or 
    end_date_field.get()=="" or 
    contact_field.get()=="" or 
    address_field=="" or 
    age_field.get()=="" ) :
        print("Enter valid details")
        messagebox.showinfo("information","Enter all the details")

    else:
        current_row=sheet.max_row
        current_column=sheet.max_column

        sheet.cell(row=current_row+1,column=1).value=name_field.get()
        sheet.cell(row=current_row+1,column=2).value=guardian_field.get()
        sheet.cell(row=current_row+1,column=3).value=join_date_field.get()
        sheet.cell(row=current_row+1,column=4).value=end_date_field.get()
        sheet.cell(row=current_row+1,column=5).value=contact_field.get()
        sheet.cell(row=current_row+1,column=6).value=address_field.get()
        sheet.cell(row=current_row+1,column=7).value=age_field.get()
        if radio_value.get()==1:
            sheet.cell(row=current_row+1,column=8).value="1 month"
        elif radio_value.get()==2:
            sheet.cell(row=current_row+1,column=8).value="3 month"
        elif radio_value.get()==3:
            sheet.cell(row=current_row+1,column=8).value="6 month"
        else:
            sheet.cell(row=current_row+1,column=8).value="12 months"
        
        worksheet.save('F:\workspace_python\Geeks For Geeks\excel1.xlsx')

        name_field.focus_set()

        clear()


def show():
    
    #new window class declaration
    check_class=Tk()
    check_class.geometry("400x400")
    status_var=StringVar()

    check_label1=Label(check_class,text="Details",font="Aerial 15 underline")
    check_label1.grid(row=0,column=1)

    search_field=Entry(check_class)

    # status_label=Label(check_class,textvariable=search_field.get())
    
     
    # book=open_workbook('F:\workspace_python\Geeks For Geeks\excel1.xlsx')
    
    # for i in book.sheets:
    #     for rowidx in range(i.nrows):
    #         row
    
def excel():
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 20
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 10

    sheet.cell(row=1,column=1).value="Name"
    sheet.cell(row=1,column=2).value="Guardian"
    sheet.cell(row=1,column=3).value="join_date"
    sheet.cell(row=1,column=4).value="end_date"
    sheet.cell(row=1,column=5).value="contact"
    sheet.cell(row=1,column=6).value="address"
    sheet.cell(row=1,column=7).value="age"
    sheet.cell(row=1,column=8).value="time"

def focus2(Event):
    guardian_field.focus_set()
def focus3(Event):
    join_date_field.focus_set()
def focus4(Event):
    end_date_field.focus_set()
def focus5(Event):
    contact_field.focus_set()
def focus6(Event):
    address_field.focus_set()
def focus7(Event):
    age_field.focus_set()


if __name__ == "__main__":
    
    main_class=Tk()
    main_class.title("Planet-X Registration form")
    main_class.geometry("600x400")
    radio_value=IntVar()
    header=Label(main_class,text="Enter Details here",font="verdana 17 underline")

    header1=Label(main_class,text="Membership for",font="Verdana 12 underline")

#Radio buttons
    time_button1=Radiobutton(main_class,text="1 Month",padx=20,variable=radio_value,value=1)
    time_button2=Radiobutton(main_class,text="3 Months",padx=20,variable=radio_value,value=2)
    time_button3=Radiobutton(main_class,text="6 Months",padx=20,variable=radio_value,value=3)
    time_button4=Radiobutton(main_class,text="12 Months",padx=20,variable=radio_value,value=4)

#Labels of UI
    name=Label(main_class,text="Full Name")
    guardian=Label(main_class,text="Father's Name/ Husband's Name")
    join_date=Label(main_class,text="Joining Date") 
    end_date=Label(main_class,text="End Date")
    contact=Label(main_class,text="Contact No.")
    address=Label(main_class,text="Address")
    age=Label(main_class,text="Age")

#Placements of the Header 
    header.grid(row=0,column=1)
    header1.grid(row=1,column=0)

#Placements of the radio buttons 
    time_button1.grid(row=2,column=0)
    time_button2.grid(row=2,column=1)
    time_button3.grid(row=3,column=0)
    time_button4.grid(row=3,column=1)

#Placements of the labels
    name.grid(row=5,column=0)
    guardian.grid(row=6,column=0)
    join_date.grid(row=7,column=0)
    end_date.grid(row=8,column=0)
    contact.grid(row=9,column=0)
    address.grid(row=10,column=0)
    age.grid(row=11,column=0)

#input Field
    name_field=Entry(main_class)
    guardian_field=Entry(main_class)
    join_date_field=Entry(main_class)
    end_date_field=Entry(main_class)
    contact_field=Entry(main_class)
    address_field=Entry(main_class)
    age_field=Entry(main_class)

#Placements of input fields
    name_field.grid(row=5,column=1,ipadx="60")
    guardian_field.grid(row=6,column=1,ipadx="60")
    join_date_field.grid(row=7,column=1,ipadx="60")
    end_date_field.grid(row=8,column=1,ipadx="60")
    contact_field.grid(row=9,column=1,ipadx="60")
    address_field.grid(row=10,column=1,ipadx="60")
    age_field.grid(row=11,column=1,ipadx="60")

#Return of the Focus
    name_field.bind("<Return>",focus2)
    guardian_field.bind("<Return>",focus3)
    join_date_field.bind("<Return>",focus4)
    end_date_field.bind("<Return>",focus5)
    contact_field.bind("<Return>",focus6)
    address_field.bind("<Return>",focus7)

#Calling excel func 
    excel()

#Buttons declarations
    submit_button=Button(main_class,text="Submit",command=insert )
    submit_button.grid(row=15,column=1)

    check_status=Button(main_class,text="Check member",command=show)
    check_status.grid(row=17,column=0)

    main_class.mainloop()
# import tkinter as tk

# root = tk.Tk()

# v = tk.IntVar()

# tk.Label(root, 
#         text="""Choose a 
# programming language:""",
#         justify = tk.LEFT,
#         padx = 20).pack()
# tk.Radiobutton(root, 
#               text="Python",
#               padx = 20, 
#               variable=v, 
#               value=1).pack(anchor=tk.W)
# tk.Radiobutton(root, 
#               text="Perl",
#               padx = 20, 
#               variable=v, 
#               value=2).pack(anchor=tk.W)

# root.mainloop()